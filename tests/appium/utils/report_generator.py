"""
OmniGuard — Excel Report Generator
===================================
Generates a styled .xlsx report after each test run.

Sheets
------
1. Summary  — pass/fail/skip counts, pass-rate %, bar chart
2. Details  — per-test row: ID, name, module, status, duration, error
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter


# ── Colours ──────────────────────────────────────────────────
CLR_PASS   = "FF22C55E"   # green-500
CLR_FAIL   = "FFEF4444"   # red-500
CLR_SKIP   = "FFF97316"   # orange-500
CLR_HEADER = "FF0B1120"   # app brand dark
CLR_ACCENT = "FF00F0FF"   # brand cyan
CLR_ROW_A  = "FF0F172A"   # alternating row dark
CLR_ROW_B  = "FF1E293B"   # alternating row lighter
CLR_WHITE  = "FFFFFFFF"


def _thin_border():
    side = Side(style="thin", color="FF334155")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font(size=11, bold=True):
    return Font(name="Calibri", bold=bold, size=size, color=CLR_WHITE)


def _cell_font(size=10, bold=False, color=CLR_WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)


# ── Helpers ──────────────────────────────────────────────────
def _parse_nodeid(nodeid: str):
    """Break pytest nodeid into (module, test_name)."""
    parts = nodeid.split("::")
    module = parts[-2] if len(parts) >= 2 else "unknown"
    name   = parts[-1] if parts else nodeid
    # strip parametrize brackets if any
    name = re.sub(r"\[.*\]", "", name)
    return module, name


# ── Summary Sheet ────────────────────────────────────────────
def _write_summary(wb: openpyxl.Workbook, results: list[dict], suite_name: str):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    total  = len(results)
    passed = sum(1 for r in results if r["outcome"] == "PASSED")
    failed = sum(1 for r in results if r["outcome"] == "FAILED")
    skipped = sum(1 for r in results if r["outcome"] == "SKIPPED")
    pass_rate = round((passed / total * 100) if total else 0, 1)
    total_dur = round(sum(r["duration"] for r in results), 2)

    # ── Title ──
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"OmniGuard — {suite_name} Test Report"
    title_cell.font      = Font(name="Calibri", bold=True, size=16, color=CLR_ACCENT)
    title_cell.fill      = _fill(CLR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Suite: {suite_name}"
    sub.font      = Font(name="Calibri", size=9, color="FF94A3B8")
    sub.fill      = _fill(CLR_HEADER)
    sub.alignment = Alignment(horizontal="center")

    # ── KPI boxes (row 4-7) ──
    kpis = [
        ("Total Tests", total,    "FF3B82F6"),
        ("Passed",      passed,   "FF22C55E"),
        ("Failed",      failed,   "FFEF4444"),
        ("Skipped",     skipped,  "FFF97316"),
        ("Pass Rate",   f"{pass_rate}%", "FF8B5CF6"),
        ("Duration",    f"{total_dur}s", "FF06B6D4"),
    ]

    col_labels = ["A", "B", "C", "D", "E", "F"]
    for idx, (label, value, color) in enumerate(kpis):
        col = col_labels[idx]
        ws.merge_cells(f"{col}4:{col}5")
        ws.merge_cells(f"{col}6:{col}7")

        lbl_cell = ws[f"{col}4"]
        lbl_cell.value     = label
        lbl_cell.fill      = _fill("FF1E293B")
        lbl_cell.font      = Font(name="Calibri", size=9, bold=True, color="FF94A3B8")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_cell = ws[f"{col}6"]
        val_cell.value     = value
        val_cell.fill      = _fill(CLR_HEADER)
        val_cell.font      = Font(name="Calibri", size=20, bold=True, color=color)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Bar chart data (hidden table rows 10-14) ──
    ws["A10"] = "Category"
    ws["B10"] = "Count"
    ws["A11"] = "Passed"
    ws["B11"] = passed
    ws["A12"] = "Failed"
    ws["B12"] = failed
    ws["A13"] = "Skipped"
    ws["B13"] = skipped

    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Test Results"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Status"
    chart.style   = 10
    chart.width   = 18
    chart.height  = 12

    data   = Reference(ws, min_col=2, min_row=10, max_row=13)
    cats   = Reference(ws, min_col=1, min_row=11, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Color individual bars
    bar_colors = ["FF22C55E", "FFEF4444", "FFF97316"]
    for i, hex_c in enumerate(bar_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = hex_c
        chart.series[0].dPt.append(pt)

    ws.add_chart(chart, "A16")

    # ── Column widths ──
    for col in col_labels:
        ws.column_dimensions[col].width = 18


# ── Details Sheet ────────────────────────────────────────────
def _write_details(wb: openpyxl.Workbook, results: list[dict]):
    ws = wb.create_sheet("Test Details", 1)
    ws.sheet_view.showGridLines = False

    headers = ["#", "Test ID", "Module", "Test Name", "Status", "Duration (s)", "Error / Notes"]
    col_widths = [5, 14, 28, 42, 10, 14, 60]

    # ── Header row ──
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = _header_font(size=11, bold=True)
        cell.fill      = _fill(CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Data rows ──
    status_fill = {
        "PASSED":  _fill(CLR_PASS),
        "FAILED":  _fill(CLR_FAIL),
        "SKIPPED": _fill(CLR_SKIP),
        "ERROR":   _fill(CLR_FAIL),
    }
    status_font = {
        "PASSED":  Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
        "FAILED":  Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
        "SKIPPED": Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
        "ERROR":   Font(name="Calibri", bold=True, size=10, color="FFFFFFFF"),
    }
    row_fills = [_fill(CLR_ROW_A), _fill(CLR_ROW_B)]

    for row_idx, result in enumerate(results, start=2):
        module, test_name = _parse_nodeid(result["nodeid"])
        status = result["outcome"]
        row_fill = row_fills[(row_idx - 2) % 2]

        row_data = [
            row_idx - 1,
            f"TC-{str(row_idx - 1).zfill(3)}",
            module,
            test_name,
            status,
            result["duration"],
            result.get("longrepr", "")[:500],  # truncate very long errors
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.font      = _cell_font()
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 2, 5, 6) else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = _thin_border()

            # Status cell special formatting
            if col_idx == 5:
                cell.fill = status_fill.get(status, row_fill)
                cell.font = status_font.get(status, _cell_font())
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 22

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"


# ── Public API ───────────────────────────────────────────────
def generate_excel_report(
    results: list[dict],
    suite_name: str = "OmniGuard",
    output_dir: Path | None = None,
) -> Path:
    """
    Generate the Excel report and save it.

    Parameters
    ----------
    results   : list of dicts with keys: nodeid, outcome, duration, longrepr
    suite_name: label shown in the report title / filename
    output_dir: directory to save the .xlsx; defaults to ./reports/

    Returns
    -------
    Path to the generated .xlsx file
    """
    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"OmniGuard_{suite_name}_Report_{timestamp}.xlsx"
    out_path  = output_dir / filename

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_summary(wb, results, suite_name)
    _write_details(wb, results)

    wb.save(out_path)
    print(f"\n📊 Excel report saved → {out_path}")
    return out_path
